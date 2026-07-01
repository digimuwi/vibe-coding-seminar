import openpyxl
import re

wb = openpyxl.load_workbook('Graupner_temp.xlsx')
ws = wb['Sacred Music']

headers = {ws.cell(1, i).value: i for i in range(1, ws.max_column + 1)}
gwv_col = headers.get('GWV')
mov_col = headers.get('MOVEMENTS')

problematic_works = []

for row_idx in range(2, ws.max_row + 1):
    gwv = ws.cell(row_idx, gwv_col).value
    mov_text = ws.cell(row_idx, mov_col).value
    
    if not gwv or not mov_text or str(mov_text).strip() == '':
        continue
    
    # Try to parse like the extractor
    mov_str = str(mov_text)
    lines = mov_str.split('\n')
    
    parsed_count = 0
    for line in lines:
        line = line.strip()
        if not line:
            continue
        pattern = r'^(\d+)\.(\w+)\s*\(([^)]+)\)'
        if re.match(pattern, line):
            parsed_count += 1
    
    # If MOVEMENTS field is not empty, but nothing can be parsed
    if mov_text and parsed_count == 0:
        problematic_works.append({
            'gwv': gwv,
            'row': row_idx,
            'text_preview': mov_str[:150]
        })

print(f'Total works checked: {ws.max_row - 1}')
print(f'Works with non-empty MOVEMENTS: {ws.max_row - 1 - 5}')

if problematic_works:
    print(f'\nWorks with MOVEMENTS text but NO parsed movements: {len(problematic_works)}')
    for w in problematic_works:
        print(f"  Row {w['row']}: GWV={w['gwv']}")
        print(f"    Preview: {w['text_preview'][:80]}...")
else:
    print('\nNo parsing problems found')

print(f'\nExpected after filtering: {ws.max_row - 1 - 5} works = 1412')
print(f'Actual in CSV: 1408 works')
print(f'Difference: {1412 - 1408} = 4 works unaccounted for')
