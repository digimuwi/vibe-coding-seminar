import openpyxl
import pandas as pd

file_path = 'Graupner GWV online.xlsx'
wb = openpyxl.load_workbook(file_path)

print('Available Sheets:', wb.sheetnames)

if 'sacred music' in wb.sheetnames:
    ws = wb['sacred music']
    print(f'\nSheet "sacred music": {ws.max_column} columns x {ws.max_row} rows')
    
    print('\nColumn Headers (first 25):')
    for i in range(1, min(26, ws.max_column + 1)):
        val = ws.cell(1, i).value
        print(f'  Col {i:2d}: {val}')
    
    print('\n\nFirst data row (Zeile 2):')
    for i in range(1, min(26, ws.max_column + 1)):
        val = ws.cell(2, i).value
        print(f'  Col {i:2d}: {val}')
    
    print('\n\nLooking for "Movements" column...')
    for i in range(1, ws.max_column + 1):
        val = ws.cell(1, i).value
        if val and 'movement' in str(val).lower():
            print(f'Found at Col {i}: {val}')
            print(f'  Example value (row 2): {ws.cell(2, i).value}')
