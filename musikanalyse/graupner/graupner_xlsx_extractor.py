#!/usr/bin/env python3
"""
graupner_xlsx_extractor.py

Extrahiert Daten aus der "Graupner GWV online.xlsx" Datei 
und erstellt eine Zwischendatei im CSV-Format.

Input: Graupner GWV online.xlsx (Sacred Music Sheet)
Output: Graupner-Saetze_strukturiert.csv

Struktur der Ausgabe:
- GWV (z.B. "1101/12")
- Satznummer (z.B. 1, 2, 3...)
- Typ (aria, choral, coro, etc.)
- Tonart (z.B. "F", "d", "B")
- Takt (z.B. "C", "3/4", "6/8")
- Besetzung (z.B. "SATB,ob(2),str,bc")
- Year
- Liturgical Occasion (German)
- Church Season
"""

import argparse
import re
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

DEFAULT_XLSX = Path(__file__).resolve().parent / "Graupner GWV online.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "Graupner-Saetze_strukturiert.csv"


def parse_movements_column(movements_text: str) -> list[dict]:
    """
    Parse the MOVEMENTS column which contains multi-line movement descriptions.
    
    Format examples:
    1.aria (B,ob(2),str,bc) - F - 3(/8) 
    2.choral (SATB,ob(2),str,bc) - F - C (allabreve) 
    1.coro+choral (SATB,ob(2),str,bc) - d - C  [combined types]
    1.coro (SATB,                               [incomplete besetzung]
    
    Returns a list of dicts with keys: num, type, besetzung, tonart, takt
    """
    if not movements_text or movements_text.strip() == "":
        return []
    
    movements = []
    lines = movements_text.split('\n')
    
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Pattern 1: number.type(s) (besetzung) - tonart - takt [optional info]
        # Handles: 1.aria, 1.coro+choral, 1.coro, etc.
        # Besetzung can be incomplete (end with comma)
        pattern = r'^(\d+)\.([a-z+]+)\s*\(([^)]*)\)\s*(?:-\s*(\w+)\s*(?:-\s*([^-()]+))?)?'
        match = re.match(pattern, line)
        
        if match:
            num = match.group(1)
            mtype = match.group(2)  # May contain +, e.g., "coro+choral"
            besetzung = match.group(3).strip()
            tonart = match.group(4).strip() if match.group(4) else ""
            takt = match.group(5).strip() if match.group(5) else ""
            
            # Clean up takt: remove trailing spaces and parenthetical info
            takt = re.sub(r'\s*\([^)]*\)\s*$', '', takt).strip()
            
            # For combined types like "coro+choral", split and process individually
            # but keep the main type as the first element
            types = mtype.split('+')
            main_type = types[0]  # Use first type as primary
            
            movements.append({
                'Satznummer': num,
                'Typ': main_type,
                'Besetzung': besetzung,
                'Tonart': tonart,
                'Takt': takt,
                'TypKombiniert': mtype  # Keep combined type info for reference
            })
        else:
            # Fallback: try to at least capture number and type
            fallback_pattern = r'^(\d+)\.([a-z+]+)'
            match = re.match(fallback_pattern, line)
            if match:
                num = match.group(1)
                mtype = match.group(2)
                types = mtype.split('+')
                
                movements.append({
                    'Satznummer': num,
                    'Typ': types[0],
                    'Besetzung': '',
                    'Tonart': '',
                    'Takt': '',
                    'TypKombiniert': mtype
                })
    
    return movements


def extract_year(gwv_str: str) -> Optional[str]:
    """
    Extract year from GWV string like '1101/12' -> '1712' or '1101/5' -> '1705'
    """
    if not gwv_str or '/' not in str(gwv_str):
        return None
    
    try:
        year_suffix = int(str(gwv_str).split('/')[-1])
        if year_suffix < 100:
            return str(1700 + year_suffix)
        else:
            return str(year_suffix)
    except:
        return None


def main():
    parser = argparse.ArgumentParser(
        description="Extrahiere Graupner-Kantaten-Daten aus XLSX und erstelle Zwischendatei CSV"
    )
    parser.add_argument(
        "--xlsx", 
        type=Path, 
        default=DEFAULT_XLSX,
        help="Pfad zur 'Graupner GWV online.xlsx' Datei"
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help="Pfad zur neuen Zwischendatei CSV"
    )
    args = parser.parse_args()
    
    # Load workbook
    print(f"Loading {args.xlsx}...")
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb["Sacred Music"]
    
    # Find column indices
    header_row = 1
    columns = {}
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(header_row, col_idx).value
        if header:
            columns[header] = col_idx
    
    print(f"Found columns: {list(columns.keys())}")
    
    # Extract required columns
    gwv_col = columns.get("GWV")
    movements_col = columns.get("MOVEMENTS")
    year_col = columns.get("YEAR")
    liturgical_d_col = columns.get("Liturgical occasion (D)")
    church_season_col = columns.get("Church Season")
    
    if not all([gwv_col, movements_col, year_col]):
        raise ValueError("Missing required columns: GWV, MOVEMENTS, YEAR")
    
    # Process rows
    total_rows = ws.max_row - 1
    print(f"Processing {total_rows} works...")
    
    rows = []
    skipped_rows = []
    
    for row_idx in range(2, ws.max_row + 1):
        gwv = ws.cell(row_idx, gwv_col).value
        movements_text = ws.cell(row_idx, movements_col).value
        year = ws.cell(row_idx, year_col).value
        liturgical_d = ws.cell(row_idx, liturgical_d_col).value if liturgical_d_col else None
        church_season = ws.cell(row_idx, church_season_col).value if church_season_col else None
        
        # Skip rows without valid MOVEMENTS data
        if not gwv:
            skipped_rows.append((row_idx, gwv, "Missing GWV"))
            continue
        
        if not movements_text or str(movements_text).strip() == "":
            skipped_rows.append((row_idx, gwv, "Empty/missing MOVEMENTS"))
            continue
        
        # Parse movements
        movements = parse_movements_column(str(movements_text))
        
        for mov in movements:
            row_dict = {
                'Werk': str(gwv),
                'Satznummer': mov['Satznummer'],
                'Typ': mov['Typ'],
                'Besetzung': mov['Besetzung'],
                'Tonart': mov.get('Tonart', ''),
                'Takt': mov.get('Takt', ''),
                'Jahr': year,
                'Liturgische_Gelegenheit': liturgical_d or '',
                'Kirchenjahr': church_season or ''
            }
            rows.append(row_dict)
    
    # Create DataFrame and save
    df = pd.DataFrame(rows)
    print(f"Extracted {len(df)} movements from {df['Werk'].nunique()} works")
    
    if skipped_rows:
        print(f"\nSkipped {len(skipped_rows)} rows due to missing data:")
        for row_idx, gwv, reason in skipped_rows:
            print(f"  Row {row_idx}: GWV={gwv} ({reason})")
    
    df.to_csv(args.output, sep=';', index=False, encoding='cp1252')
    print(f"\nSaved to {args.output}")
    
    # Print summary
    print(f"\nSummary:")
    print(f"  Total input rows: {total_rows}")
    print(f"  Processed works: {df['Werk'].nunique()}")
    print(f"  Total movements extracted: {len(df)}")
    print(f"  Movement types: {sorted(df['Typ'].unique().tolist())}")
    print(f"  Year range: {df['Jahr'].min()} - {df['Jahr'].max()}")


if __name__ == "__main__":
    main()
